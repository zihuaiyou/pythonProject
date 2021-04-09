#
# -*- coding: utf-8 -*-
#
# 这是主界面3的主程序文件
#
# 导入主程序文件所需的库

# import sys
from PyQt5.QtWidgets import *
import mainWin_3_UI
import mainWin_3_UI_child
from scipy.optimize import fsolve
from func import *
from openpyxl import Workbook

wp = " "  # 井口泵压
ph = " "  # 静液柱压力
pf = " "  # 沿程摩阻
h = " "  # 各层厚度
pc = " "  # 各层破裂压力
Q = " "  # 排量
ru = " "  # 压裂液密度
alf = " "  # 流量系数
D = " "  # 孔眼直径

# 主界面1
class myMainWindow(QMainWindow, mainWin_3_UI.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 确认输入按钮
    def inputdata(self):
        global wp, ph, pf, h, pc, Q, ru, alf, D
        wp = self.lineEdit_wp.text()
        ph = self.lineEdit_ph.text()
        pf = self.lineEdit_pf.text()
        h = self.lineEdit_h.text()
        pc = self.lineEdit_pc.text()
        Q = self.lineEdit_Q.text()
        ru = self.lineEdit_ru.text()
        alf = self.lineEdit_alf.text()
        D = self.lineEdit_D.text()
        lists = [wp, ph, pf, Q, ru]

        # 当流量系数和孔眼直径缺省时使用默认值
        if alf == '':
            alf = 0.8
        else:
            pass
        if D == '':
            D = 10
        else:
            pass

        lists_bool = []
        for content_ in lists:
            def test(content_):
                try:
                    strings = content_.split(",")
                    float_list(strings)
                    return True
                except ValueError as e:
                    return False

            if test(content_):
                lists_bool.append(True)
            else:
                lists_bool.append(False)

        # 判断参数是否输全，是否含有空格
        if False in lists_bool:
            QMessageBox.warning(self, '警告', '请完整输入所需参数并且不要加空格！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)


# 子窗口
class MychildWindow(QWidget, mainWin_3_UI_child.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 确认输出按钮
    def printdata(self):
        global pi, pm, n, N, Q1, Q2, pm1, pm2, list_new, list_3_new, list_2
        # 计算参数A,B,C
        def coefficient(*args):
            A = ((2.245e-10) * float(ru)) / ((int(args[0]) ** 2) * ((float(D) / 1000) ** 4) * (float(alf) ** 2))
            B = ((2.245e-10) * float(ru)) / ((int(args[1]) ** 2) * ((float(D) / 1000) ** 4) * (float(alf) ** 2))
            C = ((2.245e-10) * float(ru)) / ((int(args[2]) ** 2) * ((float(D) / 1000) ** 4) * (float(alf) ** 2))
            return A, B, C

        if wp and ph and pf and pc and Q and ru != " ":
            # 计算总射孔数
            list_1 = pc.split(',')  # list_1为str
            list_2 = [float(x) for x in list_1]  # list_2 为 float
            pcmax = max(list_2)
            pi = float(pcmax) - float(ph)
            pm = float(wp) - float(pi) - float(pf)
            a1 = 0.0000149 * float(Q) * (float(ru) ** 0.5)
            b1 = ((float(D) / 1000) ** 2) * float(alf) * (float(pm) ** 0.5)
            n = '%.7f' % (a1 / b1)
            N = int(float(n)) + 1
            # 文本框输出数据 总射孔数 最优各层射孔数
            self.lineEdit_n.setText(str(n))
            self.lineEdit_N.setText(str(N))

            # list_2 地层层数
            if len(list_2) == 2:
                pc1 = list_2[0]
                pc2 = list_2[1]
                # list_new 两层射孔的计算结果
                list_new = []
                # 调用外部rank函数排序
                list_rank2 = rank(2, N)
                for i in list_rank2:
                    n1 = i[0]
                    n2 = i[1]
                    A, B = coefficient(n1, n2, 1)[0], coefficient(n1, n2, 1)[1]
                    a = A - B
                    b = 2 * B * float(Q)
                    c = float(pc1) - float(pc2) - (B * (float(Q) ** 2))
                    # 求解二次方程导入quadratic
                    if quadratic(a, b, c):
                        result = quadratic(a, b, c)
                        # 判断解x1是否有效
                        if result[0] > 0 and result[0] < float(Q):
                            Q1 = float(result[0])
                            Q2 = float(Q) - Q1
                            pm1 = A * (float(Q1) ** 2)
                            pm2 = B * (float(Q2) ** 2)
                        # 解x1无效，判断解x2是否有效
                        elif result[1] > 0 and result[1] < float(Q):
                            Q1 = float(result[1])
                            Q2 = float(Q) - Q1
                            pm1 = A * (float(Q1) ** 2)
                            pm2 = B * (float(Q2) ** 2)
                        else:
                            Q1 = '无效解'
                            Q2 = '无效解'
                            pm1 = '无效解'
                            pm2 = '无效解'
                    else:
                        Q1 = '无解'
                        Q2 = '无解'
                        pm1 = '无解'
                        pm2 = '无解'
                    t1 = (n1, n2, Q1, Q2, pm1, pm2)
                    # list_new 两层射孔的计算结果
                    list_new.append(t1)
                # list_2 地层层数
            elif len(list_2) == 3:
                list_3_new = []
                pc1 = list_2[0]
                pc2 = list_2[1]
                pc3 = list_2[2]
                # 导入排列组合函数
                sequence_3 = rank(3, N)
                for l in sequence_3:
                    n1 = l[0]
                    n2 = l[1]
                    n3 = l[2]
                    A, B, C = coefficient(n1, n2, n3)[0], coefficient(n1, n2, n3)[1], coefficient(n1, n2, n3)[2]

                    # 求解多元非线性方程组
                    def func_3(i):
                        Q1, Q2, Q3 = i[0], i[1], i[2]
                        return [
                            A * (Q1 ** 2) + pc1 - B * (Q2 ** 2) - pc2,
                            A * (Q1 ** 2) + pc1 - C * (Q3 ** 2) - pc3,
                            Q1 + Q2 + Q3 - float(Q),
                        ]

                    list_r = fsolve(func_3, [0, 0, 0])
                    Q1 = float(list_r[0])
                    Q2 = float(list_r[1])
                    Q3 = float(list_r[2])
                    pm1 = A * (float(Q1) ** 2)
                    pm2 = B * (float(Q2) ** 2)
                    pm3 = C * (float(Q3) ** 2)
                    t2 = (n1, n2, n3, Q1, Q2, Q3, pm1, pm2, pm3)
                    # list_3_new 三层计算结果
                    list_3_new.append(t2)
            else:
                QMessageBox.warning(self, '警告', '目前仅支持两层或三层射孔方案 ！',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

        else:
            QMessageBox.warning(self, '警告', '数据请填写完整 ！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

    # 文件另存为按钮
    def openfile(self):
        if self.lineEdit_N.text() and self.lineEdit_n.text() != " ":
            s1 = '总射孔数：%f' % float(n)
            s2 = '最优各层射孔数：%i' % int(N)
            # list_2 地层层数
            if len(list_2) == 2:
                text = '各层射孔方案\nn1   n2   Q1(m³/min)   Q2(m³/min)   PM1(Mpa)     PM2(Mpa) \n'
                file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx);;文本(*.txt)")
                content_1 = str(s1 + '\n') + str(s2 + '\n') + text
                # 保存为txt文件
                if file[0][-3:] == 'txt':
                    f = open(file[0], encoding='utf-8', mode='a')
                    with f:
                        f.write(content_1)

                        # list_new 两层射孔的计算结果
                        for i in list_new:
                            if isinstance(i[2], float):
                                n1 = i[0]
                                n2 = i[1]
                                Q1 = '%.7f' % i[2]
                                Q2 = '%.7f' % i[3]
                                pm1 = '%.7f' % i[4]
                                pm2 = '%.7f' % i[5]

                                content = str(n1) + '     ' + str(n2) + '     ' \
                                          + str(Q1) + '     ' + str(Q2) + '     ' \
                                          + str(pm1) + '     ' + str(pm2) + '\n'
                                f.write(content)
                            else:
                                n1 = i[0]
                                n2 = i[1]
                                Q1 = i[2]
                                Q2 = i[3]
                                pm1 = i[4]
                                pm2 = i[5]

                                content = str(n1) + '     ' + str(n2) + '     ' \
                                          + str(Q1) + '     ' + str(Q2) + '     ' \
                                          + str(pm1) + '     ' + str(pm2) + '\n'
                                f.write(content)
                # 保存为xlsx文件
                else:
                    # 使用openpyxl库的Workbook()
                    wb = Workbook()
                    ws = wb.active
                    # 添加数据文件标题
                    ws['A1'] = s1
                    ws['A2'] = s2
                    ws['A3'] = '各层射孔方案'
                    list_name = ['n1', 'n2', 'Q1(m³/min)', 'Q2(m³/min)', 'PM1(Mpa)', 'PM2(Mpa)']
                    ws.append(list_name)
                    # list_new 两层射孔的计算结果
                    for i in list_new:
                        if isinstance(i[2], float):
                            n1 = i[0]
                            n2 = i[1]
                            Q1 = '%.2f' % i[2]
                            Q2 = '%.2f' % i[3]
                            pm1 = '%.2f' % i[4]
                            pm2 = '%.2f' % i[5]

                        else:
                            n1 = i[0]
                            n2 = i[1]
                            Q1 = i[2]
                            Q2 = i[3]
                            pm1 = i[4]
                            pm2 = i[5]

                        # list_data 一次计算结果
                        list_data = [n1, n2, Q1, Q2, pm1, pm2]
                        ws.append(list_data)
                    wb.save(filename=file[0])

            # list_2 地层层数
            elif len(list_2) == 3:
                text = '各层射孔方案\nn1   n2   n3   Q1(m³/min)   Q2(m³/min)   Q3(m³/min)   PM1(Mpa)     ' \
                       'PM2(Mpa)     PM3(Mpa) \n'
                file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx);;文本(*.txt)")
                content_1 = str(s1 + '\n') + str(s2 + '\n') + text
                # 保存为txt文件
                if file[0][-3:] == 'txt':
                    f = open(file[0], encoding='utf-8', mode='a')
                    with f:
                        f.write(content_1)
                        for i in list_3_new:
                            if isinstance(i[3], float):
                                n1 = i[0]
                                n2 = i[1]
                                n3 = i[2]
                                Q1 = '%.7f' % i[3]
                                Q2 = '%.7f' % i[4]
                                Q3 = '%.7f' % i[5]
                                pm1 = '%.7f' % i[6]
                                pm2 = '%.7f' % i[7]
                                pm3 = '%.7f' % i[8]

                                content = str(n1) + '     ' + str(n2) + '     ' \
                                          + str(n3) + '     ' + str(Q1) + '     ' \
                                          + str(Q2) + '     ' + str(Q3) + '     ' \
                                          + str(pm1) + '     ' + str(pm2) \
                                          + '     ' + str(pm3) + '     ' + '\n'
                                f.write(content)
                # 保存为xlsx文件
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = s1
                    ws['A2'] = s2
                    ws['A3'] = '各层射孔方案'
                    list_name = ['n1', 'n2', 'n3', 'Q1(m³/min)', 'Q2(m³/min)', 'Q3(m³/min)',
                                 'PM1(Mpa)', 'PM2(Mpa)', 'PM3(Mpa)']
                    ws.append(list_name)
                    # list_3_new 三层射孔的计算结果
                    for i in list_3_new:
                        if isinstance(i[3], float):
                            n1 = i[0]
                            n2 = i[1]
                            n3 = i[2]
                            Q1 = '%.2f' % i[3]
                            Q2 = '%.2f' % i[4]
                            Q3 = '%.2f' % i[5]
                            pm1 = '%.2f' % i[6]
                            pm2 = '%.2f' % i[7]
                            pm3 = '%.2f' % i[8]

                            # list_data 一次计算结果
                            list_data = [n1, n2, n3, Q1, Q2, Q3, pm1, pm2, pm3]
                            ws.append(list_data)
                    wb.save(filename=file[0])



        else:
            QMessageBox.warning(self, '警告', '您还未开始计算 ！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
    # 清空按钮
    def clear_all(self):
        self.lineEdit_N.setText(" ")
        self.lineEdit_n.setText(" ")


