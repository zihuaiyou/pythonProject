#
# -*- coding: utf-8 -*-
#
# 这是主界面2的主程序文件
#
# 导入主程序文件所需的库


# import sys
from PyQt5.QtWidgets import *
import mainWin_2_UI
import mainWin_2_UI_child
from func import *
from openpyxl import Workbook

# stress 各层储层的应力
# thick 各层厚度
# Q 总排量
# viscosity 粘度

# 声明全局变量前先赋值
list_param = []
list_result1 = []
list_result2 = []


# 主界面2
class myMainWindow(QMainWindow, mainWin_2_UI.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 确认输入按钮
    def inputData(self):
        global list_param
        # 各层储层的储层应力
        stress = self.lineEdit_stress.text()
        # 各层储层的储层厚度
        thick = self.lineEdit_thick.text()
        # 总排量
        Q = self.lineEdit_Q.text()
        # 粘度
        viscosity = self.lineEdit_viscosity.text()
        #
        lists = [stress, thick, Q, viscosity]
        # 含有布尔值的列表
        lists_bool = []
        # 判断参数是否输全，是否含有空格
        for content_ in lists:
            def test(content_):
                try:
                    strings = content_.split(",")
                    float_list(strings)
                    return True
                except ValueError as e:
                    return False

            if test(content_):
                lists_bool.append(True)
            else:
                lists_bool.append(False)

        # 判断参数是否输全，是否含有空格
        if False in lists_bool:
            QMessageBox.warning(self, '警告', '请完整输入所需参数并且不要加空格！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        else:
            if len(stress.split(',')) != 2 or len(thick.split(',')) != 2:
                QMessageBox.warning(self, '警告', '请输入两层储层的数据',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            else:
                # 将数据放入各自的列表(字符串)
                list_stress = stress.split(',')
                list_thick = thick.split(',')
                list_Q = Q.split(',')
                list_viscosity = viscosity.split(',')
                # 将列表数据转化为float
                # 外部自定义float_list函数
                list_stress = float_list(list_stress)
                list_thick = float_list(list_thick)
                # 导入外部自定义sumList函数,防止用户输入多层排量数据
                list_Q = sumList(float_list(list_Q))
                viscosity = float(list_viscosity[0])

                list_param = [list_stress, list_thick, list_Q, viscosity]


# 子窗口 输出结果
class myChildWindow(QWidget, mainWin_2_UI_child.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 开始计算按钮
    def calculate(self):
        # 防止用户不输入数据就开始计算
        if list_param:
            # 声明全局变量 为了使变量能在函数间传递
            global list_result1, list_result2
            # 数据初始化
            # 组合结果
            list_combination = []
            # 可以均匀扩展的组
            list_result1 = []
            # 不可以均匀扩展的组
            list_result2 = []

            # kx ky 第x,y层储层厚度
            kx = list_param[1][0]
            ky = list_param[1][1]
            # jx,jy 第x,y层储层间的储层应力
            jx = list_param[0][0]
            jy = list_param[0][1]
            # l为总排量
            l = list_param[2]
            # viscosity 粘度
            viscosity = list_param[3]
            # I1~I5 公式所需参数
            # 归一化I1~I5
            I1 = (kx - 6) / 12
            I2 = (ky - 6) / 12
            I3 = ((jy - jx) + 8) / 16
            I4 = (l - 8) / 6
            I5 = (viscosity - 5) / 395

            # 计算公式
            H11 = -1.4878 * I1 - 0.7308 * I2 - 6.9816 * I3 - 1.2318 * I4 + 1.4257 * I5 - 7.287
            H12 = 3.6286 * I1 + 5.1922 * I2 + 3.3484 * I3 - 0.9512 * I4 + 0.5138 * I5 + 3.6367
            H13 = 4.5593 * I1 - 4.3620 * I2 + 24.4709 * I3 - 77.1541 * I4 - 16.1234 * I5 - 31.5987
            H14 = 48.3322 * I1 + 16.6895 * I2 + 32.1399 * I3 - 6.4819 * I4 + 5.8499 * I5 - 7.727
            H15 = -0.7782 * I1 + 3.0725 * I2 + 3.0923 * I3 - 0.7073 * I4 + 0.4811 * I5 - 0.5468
            H16 = -1.7990 * I1 + 0.4853 * I2 + 6.8740 * I3 - 0.4623 * I4 + 3.7502 * I5 + 2.7279
            H17 = 51.0529 * I1 + 8.2303 * I2 + 32.9504 * I3 + 0.9670 * I4 - 17.0953 * I5 + 26.07
            H18 = 1.3949 * I1 - 0.4492 * I2 - 6.2934 * I3 + 0.4404 * I4 - 3.4124 * I5 - 2.7929

            list_H = [H11, H12, H13, H14, H15, H16, H17, H18]

            H21 = -1.0221 * list_H[0] + 1.1964 * list_H[1] - 0.1100 * list_H[2] - 0.8479 * list_H[
                3] - 1.600 * \
                  list_H[4] - 7.9796 * list_H[5] - 0.7832 * list_H[6] - 8.5178 * list_H[7] + 9.3214

            H22 = 1.0220 * list_H[0] - 1.1964 * list_H[1] + 0.1100 * list_H[2] + 0.8479 * list_H[
                3] + 1.600 * \
                  list_H[4] \
                  + 7.9796 * list_H[5] + 0.7832 * list_H[6] + 8.5178 * list_H[7] - 8.3214

            H21new = H21
            H22new = H22

            # 判断是否扩展

            if H21new > H22new:
                list_result1.append([1, 2])
            elif H21new < H22new:
                list_result2.append([1, 2])

            # 文本框输出结果
            if list_result1:
                result1 = ','.join(str(i) for i in list_result1)
                self.lineEdit_1.setText(str(result1))
            else:
                self.lineEdit_1.setText(" ")

            if list_result2:
                result2 = ','.join(str(i) for i in list_result2)
                self.lineEdit_2.setText(str(result2))
            else:
                self.lineEdit_2.setText(" ")
        else:
            QMessageBox.warning(self, '警告', '请完整输入参数之后再开始计算！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

    # 清空按钮
    def clear_all(self):
        self.lineEdit_1.setText(" ")
        self.lineEdit_2.setText(" ")

    # 计算结果另存为按钮
    def save_file(self):
        if list_param:
            # 将计算结果保存至excel中
            file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx)")

            # 调用openpyxl库
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '裂缝均匀扩展预测'
            list_headName = ['裂缝可以均匀扩展的组合', '裂缝不能均匀扩展的组合']

            # 添加表格标题
            ws.append(list_headName)

            # 添加计算结果
            if list_result1:
                result1 = ','.join(str(i) for i in list_result1)
                ws['A3'] = result1
            else:
                ws['A3'] = ' '

            if list_result2:
                result2 = ','.join(str(i) for i in list_result2)
                ws['B3'] = result2
            else:
                ws['B3'] = ' '

            # 保存数据文件
            wb.save(filename=file[0])
        else:
            QMessageBox.warning(self, '警告', '您还未开始计算',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)


# if __name__ == '__main__':
#     app = QApplication(sys.argv)
#     w2 = myMainWindow()
#     w2_1 = myChildWindow()
#     # 显示主界面1
#     w1.show()
#
#     # 显示子界面1
#     def show_w2():
#         w2_1.show()
#
#
#     # 绑定pushbutton_1
#     w2.pushButton_1.clicked.connect(show_w2)
#
#     app.exec_()
