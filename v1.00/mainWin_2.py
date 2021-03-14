#
# -*- coding: utf-8 -*-
#
# 这是主界面2的主程序文件
#
# 导入主程序文件所需的库


import sys
from PyQt5.QtWidgets import *
import mainWin_2_UI
import mainWin_2_UI_child
from func import *
import itertools
from openpyxl import Workbook
import math

# num 裂缝沟通组数量
# stress 裂缝沟通组间的储层应力
# thick 裂缝沟通组厚度
# Q 各组排量
# viscosity 粘度

# 声明全局变量前先赋值
list_param = []
list_result1 = []
list_result2 = []


# 主界面2
class myMainWindow(QMainWindow, mainWin_2_UI.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 确认输入按钮
    def inputData(self):
        global list_param
        # 裂缝沟通组数量
        num = self.lineEdit_num.text()
        # 裂缝沟通组间的储层应力
        stress = self.lineEdit_stress.text()
        # 裂缝沟通组厚度
        thick = self.lineEdit_thick.text()
        # 各组排量
        Q = self.lineEdit_Q.text()
        # 粘度
        viscosity = self.lineEdit_viscosity.text()
        # 将数据放入各自的列表(字符串)
        list_stress = stress.split(',')
        list_thick = thick.split(',')
        list_Q = Q.split(',')
        list_viscosity = viscosity.split(',')
        # 将列表数据转化为float
        # 外部自定义float_list函数
        list_stress = float_list(list_stress)
        list_thick = float_list(list_thick)
        list_Q = float_list(list_Q)
        viscosity = float(list_viscosity[0])
        num = int(num)

        list_param = [num, list_stress, list_thick, list_Q, viscosity]


# 子窗口 输出结果
class myChildWindow(QWidget, mainWin_2_UI_child.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 开始计算按钮
    def calculate(self):
        # 声明全局变量 为了使变量能在函数间传递
        global list_result1, list_result2
        # 数据初始化
        # 组合结果
        list_combination = []
        # 可以均匀扩展的组
        list_result1 = []
        # 不可以均匀扩展的组
        list_result2 = []
        # 为补获异常设置的新数组
        list_new = []

        # 导入排列组合函数
        num = list_param[0]
        # list_zuhe 组合结果
        list_zuhe = itertools.combinations(range(1, num + 1), 2)
        # 确保上下层关系
        for n in list_zuhe:
            if n[0] < n[1]:
                list_combination.append(n)

        # 判断每种组合能否均匀扩展
        for i in list_combination:
            # x,y 为某一种组合情况
            # x 为上层,y为下层
            x = i[0]
            y = i[1]
            # kx ky 第x,y组裂缝沟通组厚度
            kx = list_param[2][x - 1]
            ky = list_param[2][y - 1]
            # jx,jy 第x,y组裂缝沟通组间的储层应力
            jx = list_param[1][x - 1]
            jy = list_param[1][y - 1]
            # lx,ly 第x,y组排量
            lx = list_param[3][x - 1]
            ly = list_param[3][y - 1]
            # viscosity 粘度
            viscosity = list_param[4]
            # I1~I5 公式所需参数
            I1 = kx
            I2 = ky
            I3 = jy - jx
            I4 = lx + ly
            I5 = viscosity

            # 计算公式
            H11 = -1.4878 * I1 - 0.7308 * I2 - 6.9816 * I3 - 1.2318 * I4 + 1.4257 * I5 - 7.287
            H12 = 3.6286 * I1 + 5.1922 * I2 + 3.3484 * I3 - 0.9512 * I4 + 0.5138 * I5 + 3.6367
            H13 = 4.5593 * I1 - 4.3620 * I2 + 24.4709 * I3 - 77.1541 * I4 - 16.1234 * I5 - 31.5987
            H14 = 48.3322 * I1 + 16.6895 * I2 + 32.1399 * I3 - 6.4819 * I4 + 5.8499 * I5 - 7.727
            H15 = -0.7782 * I1 + 3.0725 * I2 + 3.0923 * I3 - 0.7073 * I4 + 0.4811 * I5 - 0.5468
            H16 = -1.7990 * I1 + 0.4853 * I2 + 6.8740 * I3 - 0.4623 * I4 + 3.7502 * I5 + 2.7279
            H17 = 51.0529 * I1 + 8.2303 * I2 + 32.9504 * I3 + 0.9670 * I4 - 17.0953 * I5 + 26.07
            H18 = 1.3949 * I1 - 0.4492 * I2 - 6.2934 * I3 + 0.4404 * I4 - 3.4124 * I5 - 2.7929

            list_H = [H11, H12, H13, H14, H15, H16, H17, H18]
            # 处理计算值过大的异常
            for j in list_H:
                try:
                    result = 1 / (1 + math.e ** (-j))
                except OverflowError:
                    QMessageBox.warning(self, '警告', '方程无解 请检查所输入参数是否有误 ！',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)


                else:
                    list_new.append(result)

            # H11new = 1 / (1 + math.e ** (-H11))
            # H12new = 1 / (1 + math.e ** (-H12))
            # H13new = 1 / (1 + math.e ** (-H13))
            # H14new = 1 / (1 + math.e ** (-H14))
            # H15new = 1 / (1 + math.e ** (-H15))
            # H16new = 1 / (1 + math.e ** (-H16))
            # H17new = 1 / (1 + math.e ** (-H17))
            # H18new = 1 / (1 + math.e ** (-H18))

            H21 = -1.0221 * list_new[0] + 1.1964 * list_new[1] - 0.1100 * list_new[2] - 0.8479 * list_new[3] - 1.600 * \
                  list_new[4] - 7.9796 * list_new[5] - 0.7832 * list_new[6] - 8.5178 * list_new[7] + 9.3214

            H22 = 1.0220 * list_new[0] - 1.1964 * list_new[1] + 0.1100 * list_new[2] + 0.8479 * list_new[3] + 1.600 * \
                  list_new[4] \
                  + 7.9796 * list_new[5] + 0.7832 * list_new[6] + 8.5178 * list_new[7] - 8.3214

            H21new = H21
            H22new = H22

            # 判断是否扩展

            if H21new > H22new:
                list_result1.append([x, y])
            elif H21new < H22new:
                list_result2.append([x, y])

        print(list_result1, list_result2)

        # 文本框输出结果
        if list_result1:
            result1 = ','.join(str(i) for i in list_result1)
            self.lineEdit_1.setText(str(result1))
        else:
            self.lineEdit_1.setText(" ")

        if list_result2:
            result2 = ','.join(str(i) for i in list_result2)
            self.lineEdit_1.setText(str(result2))
        else:
            self.lineEdit_1.setText(" ")

    # 清空按钮
    def clear_all(self):
        self.lineEdit_1.setText(" ")
        self.lineEdit_2.setText(" ")

    # 计算结果另存为按钮
    def save_file(self):
        # 将计算结果保存至excel中
        file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx)")

        # 调用openpyxl库
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '裂缝沟通组间裂缝是否均匀扩展判断'
        list_headName = ['裂缝可以均匀扩展的组合', '裂缝不能均匀扩展的组合']

        # 添加表格标题
        ws.append(list_headName)

        # 添加计算结果
        if list_result1:
            result1 = ','.join(str(i) for i in list_result1)
            ws['A3'] = result1
        else:
            ws['A3'] = ' '

        if list_result2:
            result2 = ','.join(str(i) for i in list_result2)
            ws['B3'] = result2
        else:
            ws['B3'] = ' '

        # 保存数据文件
        wb.save(filename=file[0])


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w1 = myMainWindow()
    w2 = myChildWindow()
    # 显示主界面1
    w1.show()


    # 显示子界面1
    def show_w2():
        w2.show()


    # 绑定pushbutton_1
    w1.pushButton_1.clicked.connect(show_w2)

    app.exec_()
