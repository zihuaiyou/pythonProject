#
# -*- coding: utf-8 -*-
#
# 这是主界面1的主程序文件
#
# 导入主程序文件所需的库

import sys
from PyQt5.QtWidgets import *
import mainWin_1_UI
import mainWin_1_UI_child_1
import mainWin_1_UI_child_2
from func import *
from openpyxl import Workbook

# num 储层层数
# stress_diff 储隔层应力差
# reservoir_thick 储层厚度
# upper_thick 上层隔层厚度
# lower_thick 下层隔层厚度
# Q 各层排量
# scale 各层液体规模
# viscosity 粘度

# 转换为全局变量前要提前声明
list_stress_diff = []
list_reservoir_thick = []
list_upper_thick = []
list_lower_thick = []
list_Q = []
list_scale = []

# 提前给全局变量赋值
viscosity = ''
num = 0
list_param = []
list_result = []


# 主界面1
class myMainWindow(QMainWindow, mainWin_1_UI.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 确认输入按钮
    def inputData(self):
        global list_stress_diff, list_reservoir_thick, list_upper_thick, \
            list_lower_thick, list_Q, list_scale, viscosity, num
        # 转为全局变量 让子窗口能获取变量
        global list_param, num, viscosity
        # 获取输入的参数(字符串形式且含逗号)

        # 储层层数
        num = self.lineEdit_num.text()
        # 储隔层应力差
        stress_diff = self.lineEdit_stress_diff.text()
        # 储层厚度
        reservoir_thick = self.lineEdit_reservoir_thick.text()
        # 上层隔层厚度
        upper_thick = self.lineEdit_upper_thick.text()
        # 下层隔层厚度
        lower_thick = self.lineEdit_lower_thick.text()
        # 各层排量
        Q = self.lineEdit_Q.text()
        # 各层液体规模
        scale = self.lineEdit_scale.text()
        # 粘度
        viscosity = self.lineEdit_viscosity.text()
        #
        lists = [num, stress_diff, reservoir_thick, upper_thick, lower_thick, Q, scale, viscosity]
        lists_bool = []
        # 判断参数是否输全，是否含有空格
        for content_ in lists:
            def test(content_):
                try:
                    strings = content_.split(",")
                    float_list(strings)
                    return True
                except ValueError as e:
                    return False

            if test(content_):
                lists_bool.append(True)
            else:
                lists_bool.append(False)

        # 判断参数是否输全，是否含有空格
        if False in lists_bool:
            QMessageBox.warning(self, '警告', '请完整输入所需参数并且不要加空格！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        else:
            # 将数据放入各自的列表(字符串)
            list_stress_diff = stress_diff.split(',')
            list_reservoir_thick = reservoir_thick.split(',')
            list_upper_thick = upper_thick.split(',')
            list_lower_thick = lower_thick.split(',')
            list_Q = Q.split(',')
            list_scale = scale.split(',')
            list_viscosity = viscosity.split(',')
            # 将列表数据转化为float
            # 外部自定义float_list函数
            list_stress_diff = float_list(list_stress_diff)
            list_reservoir_thick = float_list(list_reservoir_thick)
            list_upper_thick = float_list(list_upper_thick)
            list_lower_thick = float_list(list_lower_thick)
            list_Q = float_list(list_Q)
            list_scale = float_list(list_scale)

            viscosity = float(list_viscosity[0])
            num = int(num)
            list_param = [list_stress_diff, list_reservoir_thick, list_upper_thick,
                          list_lower_thick, list_Q, list_scale]

            # 判断所输入层数是否和其他参数匹配！
            list_bool_2 = []
            for i in list_param:
                if len(i) == num:
                    list_bool_2.append(True)
                else:
                    list_bool_2.append(False)

            if False in list_bool_2:
                QMessageBox.warning(self, '警告', '所输入层数要和其他参数匹配！',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

# 子界面1 各层人工裂缝缝高预测
class myChildWindow1(QWidget, mainWin_1_UI_child_1.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 让表格内容不可编辑
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)

    # 开始计算按钮
    def calculation(self):
        global list_param
        # 防止用户不输入数据就开始计算
        if list_param:
            list_interlayer_thick = []
            list_HF = []
            # 计算隔层厚度
            for i, j in zip(list_param[2], list_param[3]):
                l = (i + j) / 2
                list_interlayer_thick.append(l)

            # 计算人工裂缝缝高HF
            for x in range(num):
                # 各层人工裂缝缝高预测公式
                # HF 人工裂缝缝高
                HF = -1.3517 * list_param[0][x] + 3.0332 * list_param[1][x] \
                     - 0.5259 * list_interlayer_thick[x] + 2.3170 * list_param[4][x] + 0.0496 \
                     * viscosity + 0.0360 * list_param[5][x]

                # 结果保留两位小数
                HF = round(HF, 2)
                list_HF.append(HF)

            list_param.insert(0, list_HF)

            # 向tableWidget中添加计算结果
            self.tableWidget.setRowCount(num)
            for x in range(num):
                for y in range(7):
                    content = str(list_param[y][x])
                    self.tableWidget.setItem(x, y, QTableWidgetItem(content))
        else:
            QMessageBox.warning(self, '警告', '请完整输入参数之后再开始计算！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

    # 计算结果另存为按钮
    def saveData(self):
        # 防止用户不输入数据就保存文件
        if list_param:
            list_count = []

            # 将计算结果保存至excel中
            file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx)")

            # 调用openpyxl库
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '各层人工裂缝缝高预测'
            list_headName = ['储层序号', '人工裂缝缝高(m)', '储隔层应力差(MPa)', '储层厚度(m)',
                             '上层隔层厚度(m)', '下层隔层厚度(m)', '排量(m³/min)', '液体规模(m³)']

            # 添加表格标题
            ws.append(list_headName)

            # 向表格第一列添加字符'第X层'
            for i in range(num):
                str = '第%s层' % (i + 1)
                list_count.append(str)

            # 增加序列 "第x层"
            list_param.insert(0, list_count)

            # 按行添加至excel 所以需用zip转置list
            result = list(zip(*list_param))
            for i in range(num):
                ws.append(list(result[i]))
            if file[0]:
                wb.save(filename=file[0])
        else:
            QMessageBox.warning(self, '警告', '您还未开始计算',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)


# 子界面2 裂缝纵向沟通预判
class myChildWindow2(QWidget, mainWin_1_UI_child_2.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 开始计算按钮
    def calculate(self):
        global list_result
        # 隔层厚度
        list_interlayer_thick_2 = []
        # 人工裂缝高度
        list_HF_2 = []

        # 不穿层
        list_1 = []
        # 穿上层
        list_2 = []
        # 穿下层
        list_3 = []
        # 穿两层
        list_4 = []
        # 防止用户不输入数据就开始计算
        if list_param:
            # 计算隔层厚度
            for i, j in zip(list_param[2], list_param[3]):
                l = (i + j) / 2
                list_interlayer_thick_2.append(l)

            # 在进行裂缝纵向分组时,组数必须超过一组
            if int(num) == 1:
                QMessageBox.warning(self, '警告', '组数必须超过一组',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            else:
                # 计算人工裂缝缝高HF
                for x in range(num):
                    # 各层人工裂缝缝高预测公式
                    # HF 人工裂缝缝高
                    HF = -1.3517 * list_stress_diff[x] + 3.0332 * list_reservoir_thick[x] \
                         - 0.5259 * list_interlayer_thick_2[x] + 2.3170 * list_Q[x] + 0.0496 \
                         * viscosity + 0.0360 * list_scale[x]

                    # 结果保留两位小数
                    HF = round(HF, 2)
                    list_HF_2.append(HF)

                # 判断各储层的人工裂缝穿层情况
                # 第一层和最后一层储层要单独考虑
                # 第一层储层 只有穿下层和不穿层情况
                for x in range(num):
                    if x == 0:
                        func2 = ((list_HF_2[x] - list_reservoir_thick[x]) / 2 + (list_HF_2[x + 1] -
                                                                                 list_reservoir_thick[x + 1]) / 2)
                        if func2 < list_lower_thick[x]:
                            # 不穿层
                            list_1.append(x + 1)
                        elif func2 > list_lower_thick[x]:
                            # 穿下层
                            list_3.append(x + 1)
                    # 最后一层储层 只有穿上层和不穿层情况
                    elif x == num - 1:
                        func1 = ((list_HF_2[x] - list_reservoir_thick[x]) / 2 + (list_HF_2[x - 1] -
                                                                                 list_reservoir_thick[x - 1]) / 2)
                        if func1 < list_upper_thick[x]:
                            # 不穿层
                            list_1.append(x + 1)
                        elif func1 > list_upper_thick[x]:
                            # 穿上层
                            list_2.append(x + 1)
                    # 中间储层 四种情况
                    elif 0 < x < num - 1:
                        # func1,func2 公式1,公式2
                        func1 = ((list_HF_2[x] - list_reservoir_thick[x]) / 2 + (list_HF_2[x - 1] -
                                                                                 list_reservoir_thick[x - 1]) / 2)
                        func2 = ((list_HF_2[x] - list_reservoir_thick[x]) / 2 + (list_HF_2[x + 1] -
                                                                                 list_reservoir_thick[x + 1]) / 2)
                        if func1 < list_upper_thick[x] and func2 < list_lower_thick[x]:
                            # 不穿层
                            list_1.append(x + 1)
                        elif func1 > list_upper_thick[x] and func2 < list_lower_thick[x]:
                            # 穿上层
                            list_2.append(x + 1)
                        elif func1 < list_upper_thick[x] and func2 > list_lower_thick[x]:
                            # 穿下层
                            list_3.append(x + 1)
                        elif func1 > list_upper_thick[x] and func2 > list_lower_thick[x]:
                            # 穿两层
                            list_4.append(x + 1)

                # 裂缝沟通分组
                # 裂缝纵向沟通组 [组号,储层序号]
                # 储层1放入裂缝沟通组1
                list_fenzu = [[1, 1]]
                m = 1
                for i in range(num):
                    # j 储层序号
                    j = i + 1
                    # 从第二层储层开始
                    if j > 1:
                        # 判断第j层储层穿层情况
                        if j in list_1:
                            if j - 1 in list_1 or j - 1 in list_2:
                                m += 1
                                list_fenzu.append([m, j])
                            elif j - 1 in list_3 or j - 1 in list_4:
                                list_fenzu.append([m, j])
                        elif j in list_2:
                            if j - 1 in list_1 or j - 1 in list_2:
                                m += 1
                                list_fenzu.append([m, j])
                            elif j - 1 in list_3 or j - 1 in list_4:
                                list_fenzu.append([m, j])
                        elif j in list_3:
                            if j - 1 in list_1 or j - 1 in list_2:
                                m += 1
                                list_fenzu.append([m, j])
                            elif j - 1 in list_3 or j - 1 in list_4:
                                list_fenzu.append([m, j])
                        elif j in list_4:
                            list_fenzu.append([m, j])

                # 向tableWidget中添加计算结果
                set1 = set()
                # 确定组数number
                for i in list_fenzu:
                    set1.add(i[0])
                number = max(set1)
                self.tableWidget.setRowCount(number)

                # 向tableWidget中添加第一列 组数
                list_fenzu2 = []
                for x in list_fenzu:
                    self.tableWidget.setItem(x[0] - 1, 0, QTableWidgetItem(str(x[0])))
                    list_fenzu2.append(str(x[1]))

                # 表示每一组有多少层储层的列表
                count_list = []
                # 表示储层在不同组分布的列表
                list_result = []
                # 转置列表
                result = list(zip(*list_fenzu))
                for i in range(number):
                    count_list.append(countX(result[0], i + 1))
                #
                x1 = 0
                for i in count_list:
                    y1 = x1 + int(i)
                    list_result.append(result[1][x1:y1])
                    x1 = y1

                # 向tableWidget中添加第二列 储层在不同组的分布
                for index, i in enumerate(list_result):
                    new_result = ",".join(str_list(i))
                    self.tableWidget.setItem(index, 1, QTableWidgetItem(str(new_result)))
        else:
            QMessageBox.warning(self, '警告', '请完整输入参数之后再开始计算！',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

    # 计算结果另存为按钮
    def saveData(self):
        # 防止用户不输入数据就保存文件
        if list_param:
            # 将计算结果保存至excel中
            file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx)")

            # 调用openpyxl库
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '裂缝纵向沟通预判'
            list_headName = ['裂缝纵向沟通组', '储层序号']
            # 添加表格标题
            ws.append(list_headName)
            # 将计算结果添加至表格中
            for index, i in enumerate(list_result):
                # string 第一列序号
                string = "A%d" % (index + 3)
                ws[string] = index + 1
                # content 第二列序号
                content = "B%d" % (index + 3)
                new_result = ",".join(str_list(i))
                ws[content] = new_result
            if file[0]:
                wb.save(filename=file[0])
        else:
            QMessageBox.warning(self, '警告', '您还未开始计算',
                                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w1 = myMainWindow()
    w1_1 = myChildWindow1()
    w1_2 = myChildWindow2()
    # 显示主界面1
    w1.show()

    # 显示子界面1
    def show_w1_1():
        w1_1.show()


    # 显示子界面2
    def show_w1_2():
        w1_2.show()


    # 绑定pushbutton
    w1.pushButton_1.clicked.connect(show_w1_1)
    w1.pushButton_2.clicked.connect(show_w1_2)

    app.exec_()
